#!/usr/bin/env python3
"""Append this Pi's running time to a spreadsheet, for battery endurance tests.

What it does
    Every few seconds, writes how long the script has been running into an
    .xlsx. Left on a battery-powered Pi, the last row tells you how long the
    battery lasted.

Hardware
    Any Pi.

Wiring
    None.

Run on a Pi
    python3 recipes/system/log_uptime_to_xlsx.py

Run on your Mac
    python3 recipes/system/log_uptime_to_xlsx.py
    (creates the workbook in the current directory)

Copy into a project
    Saving the whole workbook every interval is the point - the power can be
    cut at any moment, and an unsaved workbook tells you nothing. For anything
    long-running prefer a CSV append, which is cheaper and crash-safe.

requires: openpyxl
"""
import os
import time

from openpyxl import Workbook, load_workbook

# --- tunables -------------------------------------------------------------
LOGBOOK = os.path.join(os.getcwd(), "battery_test.xlsx")
INTERVAL_SECONDS = 5
# --------------------------------------------------------------------------


def open_logbook(path=LOGBOOK):
    """Load the workbook, creating it with a header row if absent."""
    if os.path.exists(path):
        workbook = load_workbook(path)
        return workbook, workbook.active
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(("timestamp", "uptime_hours"))
    workbook.save(path)
    return workbook, sheet


if __name__ == "__main__":
    started = time.time()
    hours = 0.0
    workbook, sheet = open_logbook()
    try:
        while True:
            time.sleep(INTERVAL_SECONDS)
            hours = (time.time() - started) / 3600
            sheet.append((time.strftime("%Y-%m-%dT%H:%M:%S"), hours))
            workbook.save(LOGBOOK)
            print("logged {:.4f} hours".format(hours))
    except KeyboardInterrupt:
        pass
    finally:
        # Seeded above, so an interrupt in the first interval cannot NameError.
        print("ran for {:.4f} hours".format(hours))
